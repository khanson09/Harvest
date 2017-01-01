from tkinter import *
from tkinter.ttk import *
import tkinter.ttk as ttk
import tkinter  as tk
import json, os, subprocess
import labels
from reportlab.graphics import shapes
from reportlab.pdfbase.pdfmetrics import stringWidth
import win32api, win32print, tempfile
import openpyxl as pyxl

#initail imports to make this behemoth work
############################################################################################################################################################################################################################################################

gather = open('fgn.json', 'r')
VisAiddictionary = json.loads(gather.read())
gather.close()

collect = open('lghtgid.json', 'r')
lghtguidedictionary = json.loads(collect.read())
collect.close()

harvest = open('bom.json', 'r')
bomdictionary = json.loads(harvest.read())
harvest.close()

#dictionaries
############################################################################################################################################################################################################################################################

# dictionary = {'42554919': 'Bolt 2018 2', '42554921': 'Bolt 2018 1','84163260': 'Acadia 1', '84163262': 'Acadia 3', '84163263': 'Acadia 1', '84163264': 'Acadia 3', '84189056': 'Acadia 2', '84189055': 'Acadia 4', '22878018': 'Camaro 1', '22878019': 'Camaro 1', '84076840': 'CT6 1','84076841':'CT6 1', '84076842': 'CT6 1', '23475843': 'Colorado 2', '23475842': 'Colorado 1', '23389053': 'Colorado 1', '84063810': 'Colorado 3','84063811':'Colorado 3','84063812': 'Colorado 3',}

############################################################################################################################################################################################################################################################

class Application(tk.Frame):

	def __init__(self, master = None):

		super().__init__(master)
		self.grid()
		root.title('Harvest')
		root.config(background = 'black')
		self.frame = Frame(self.master)

		self.notebook = ttk.Notebook(root)
		self.notebook.grid()

		SearchFGN = tk.Frame(root)
		SearchFGN.grid()
		SearchFGN.config(background = 'Black')
		addvisaid = ttk.Frame(root)
		addvisaid.grid()

		self.notebook.add(SearchFGN, text = 'Search for VisAid', state = 'normal')
		self.notebook.add(addvisaid, text = 'Add VisAid', state = 'normal')
		self.notebook.hide(addvisaid)

#Setting up the notebook that is the display of harvest
############################################################################################################################################################################################################################################################

		self.directions = Label(SearchFGN) 
		self.directions.grid(row = 0, column = 0, padx = 200, pady = 5) 
		self.directions['text'] = 'Enter Finished Good Number Here:'
		self.directions.config(font = ('Helvetica', 20), background = 'black', foreground = 'white')

		self.textentry = Entry(SearchFGN) #This is the text field where you enter in FGN's
		self.textentry.grid(row = 0, column = 1, padx = 50, pady = 5) #This affects the placement of the Text Field

		self.buildNumber = Label(SearchFGN)
		self.buildNumber.grid(row = 1, column = 0, padx = 10)
		self.buildNumber['text'] = 'How many parts are you building?'
		self.buildNumber.config(font = ('Helvetica', 20), background = 'black', foreground = 'white')

		self.NoB = Entry(SearchFGN)
		self.NoB.grid(row = 1, column = 1, pady = 5)

		self.pickbtn = Button(SearchFGN, text = 'Get VisAid', command = self.searchfgn)
		self.pickbtn.grid(row = 2, column = 1, pady = 10)
		self.pickbtn.config(width = 20)

		self.VisAidbtn = Button(SearchFGN, text = 'Add New VisAid', command = self.pass_entry)
		self.VisAidbtn.grid(row = 2, column = 0, pady = 10)

		self.errormsg = Label(SearchFGN)
		self.errormsg.grid(row = 4, column = 0, pady = 5, columnspan = 2)
		self.errormsg.config(font = ('Helvetica', 20), background = 'black')

#Setting up the main tab that is initaly displayed when Harvest is launched
############################################################################################################################################################################################################################################################

	def searchfgn(self):

		FGN = self.textentry.get()
		num_to_build = self.NoB.get()
		wb = pyxl.load_workbook('F:\\Harvest\\Trial BoM\'s\\2016-17_CAMARO_FINAL_BoM_RFQ_317973.xlsx')
		

		if FGN == '':

			self.errormsg['text'] = 'No Entry Made Please Input a Finished Good Number'
			self.errormsg.config(background = 'red', foreground = 'white')

		elif num_to_build == '':
		
			self.errormsg['text'] = 'No Build Number Specified, Please Enter Number to Build'
			self.errormsg.config(background = 'red', foreground = 'white')

		elif FGN in VisAiddictionary:
		
			pickList = tk.Toplevel(root)
			pickList.config(background = 'black')

			self.label = Label(pickList) #This Creates the label for the out put of the visual aid
			self.label.grid(row = 0, column = 0, columnspan = 3, pady = 10, padx = 10) #Assigns placment
			self.label.config(font = ('Helvetica', 20), background = 'light green')

			display = "Visual Aid for Finished Good Number {}: {}".format(FGN, VisAiddictionary[FGN])

			self.label['text'] = display

			self.msg = Label(pickList)
			self.msg.grid(row = 2, column = 0, columnspan = 3, pady = 10)
			self.msg.config(font = ('Helvetica', 18), background = 'purple', foreground = 'Pink')
			self.msg['text'] = 'For Quantity of 0 corresponds to Adhesive check with given BoM if not the same PN contact Kyle.'

			self.lghtguide = Label(pickList)
			self.lghtguide.grid(row = 10, column = 0, columnspan = 3, pady = 10)
			self.lghtguide.config(font = ('Helvetica', 20), background = 'Blue', foreground = 'white')
			self.lghtguide['text'] = 'LightGuide Program for Finished Good Number {}: {}'.format(FGN, lghtguidedictionary[FGN])

			self.errormsg['text'] = ''
			self.errormsg.config(background = 'black')

			fgn_header = ['Quantity', 'Part Number']
			quantity_list = []
			quantity = []
			part_list = []

			ws = wb[FGN]

			value = 0
 
			for row in ws.iter_rows(min_col = 1, max_col = 1, min_row = 1, max_row = 50):
				for cell in row:
					if cell.value is None:
						last_row = value
					else:
						value = value + 1

 
			for row in ws.iter_rows(min_col = 3, max_col = 3, min_row = 3, max_row = last_row):
				for cell in row:
					quantity.insert(0, int(cell.value)*int(num_to_build))
 
 
			for row in ws.iter_rows(min_col = 2, max_col = 2, min_row = 3, max_row = last_row):
				for cell in row:
					part_list.insert(0, int(cell.value))

			container = ttk.Frame()
			container.grid()

			quantity_list = zip(quantity, part_list)

				# create a treeview with dual scrollbars
			self.tree = ttk.Treeview(pickList, columns = fgn_header, show = "headings")
			vsb = ttk.Scrollbar(pickList, orient = "vertical", command = self.tree.yview)


			self.tree.configure(yscrollcommand = vsb.set)
			self.tree.grid(column=0, row=1, columnspan = 3, sticky='nsew')

			vsb.grid(column=3, row=1, sticky='ns')


			container.grid_columnconfigure(0, weight=1)
			container.grid_rowconfigure(0, weight=1)



			for col in fgn_header:
				self.tree.heading(col, text=col.title())
					# adjust the column's width to the header string

			for item in quantity_list:
				self.tree.insert('', 'end', values=item)
					# adjust column's width if necessary to fit each value


			self.drawpdf()

		else:

			self.errormsg['text'] = 'Finished Good Number: {} not found, try inputting the Finished Good Number again.'.format(FGN)
			self.errormsg.config(background = 'red', foreground = 'white')


#The search function that looks for a FGN and the corresponding Visual Aid
############################################################################################################################################################################################################################################################

	def test(self):

		print('it worked')

#test code
############################################################################################################################################################################################################################################################

	def add(self):

		FGN = self.FGNEntry.get()
		VISAID = self.VisaidEntry.get()

		if FGN in VisAiddictionary:

			display = 'Visual Aid for Finished Good Number Already exits as: {}'.format(VisAiddictionary[FGN])
			self.job.config(font = ('Helvetica', 20), background = 'red')

		elif FGN == '':

			display = 'No FGN entry made Please input an entry.'
			self.job.config(font = ('Helvetica', 20), background = 'red')

		elif VISAID == '':

			display = 'No Visual Aid entry made Please input an entry.'
			self.job.config(font = ('Helvetica', 20), background = 'red')

		else:

			gather = open('fgn.json', 'w')
			VisAiddictionary[self.FGNEntry.get()] = self.VisaidEntry.get()
			gather.write(json.dumps(VisAiddictionary))
			gather.close()
			display = 'Visual Aid: {} added as: {}'.format(FGN, VisAiddictionary[FGN])
			self.job.config(font = ('Helvetica', 20), background = 'light green')

		self.job['text'] = display

#Code that adds new visual aids to the referance dictionary
############################################################################################################################################################################################################################################################

	def show(self):

		addvisaid = ttk.Frame(root)
		addvisaid.grid()
		self.notebook.add(addvisaid, text = 'Add VisAid', state = 'normal')

		self.btn = Button(addvisaid, text = 'Add New VisAid', command = self.add)
		self.btn.grid(row = 0, column = 2, rowspan = 2)

		self.newFGN = Label(addvisaid)
		self.newFGN.grid(row = 0, column = 0, padx = 35, pady = 5)
		self.newFGN['text'] = 'New FGN:'
		self.newFGN.config(font = ('Helvetica', 20))

		self.newVisaid = Label(addvisaid)
		self.newVisaid.grid(row = 1, column = 0, padx = 35, pady = 5)
		self.newVisaid['text'] = 'New VisAid:'
		self.newVisaid.config(font = ('Helvetica', 20))

		self.VisaidEntry = Entry(addvisaid)
		self.VisaidEntry['textvariable'] = self.VisaidEntry
		self.VisaidEntry.grid(row = 1, column = 1, padx = 35, pady = 5)

		self.FGNEntry = Entry(addvisaid)
		self.FGNEntry['textvariable'] = self.FGNEntry
		self.FGNEntry.grid(row = 0, column = 1, padx = 35, pady = 5)

		self.job = Label(addvisaid)
		self.job.grid(row = 5, column = 0, pady = 5, columnspan = 3)

		self.newlght = Label(addvisaid)
		self.newlght.grid(row = 3, column = 0, padx = 35, pady = 5)
		self.newlght['text'] = 'New LightGuide:'
		self.newlght.config(font = ('Helvetica', 20))

		self.lghtEntry = Entry(addvisaid)
		self.lghtEntry['textvariable'] = self.lghtEntry
		self.lghtEntry.grid(row = 3, column = 1, padx = 35, pady = 5)

		self.btn2 = Button(addvisaid, text = 'Add New LightGuide', command = self.addlight)
		self.btn2.grid(row = 3, column = 2)		

		self.newBoM = Label(addvisaid)
		self.newBoM.grid(row = 4, column = 0, padx = 35, pady = 5)
		self.newBoM['text'] = 'New BoM:'
		self.newBoM.config(font = ('Helvetica', 20))

		self.BoMEntry = Entry(addvisaid)
		self.BoMEntry['textvariable'] = self.BoMEntry
		self.BoMEntry.grid(row = 4, column = 1, padx = 35, pady = 5)

		self.btn3 = Button(addvisaid, text = 'Add New BoM', command = self.addBoM)
		self.btn3.grid(row = 4, column = 2)	

#What makes the add new visaid tab display
############################################################################################################################################################################################################################################################

	def drawpdf(self):

		specs = labels.Specification(76.2, 50.8, 1, 1, 63.5, 38.1, corner_radius = 0)
		Number_building = int(self.NoB.get())                   
		FGN = self.textentry.get()
		wb = pyxl.load_workbook('F:\\Harvest\\Trial BoM\'s\\2016-17_CAMARO_FINAL_BoM_RFQ_317973.xlsx')
		ws = wb[FGN]
		PanelNum = ws['B3'].value
 
		def write_sticker(label, width, height, name):
 
			label.add(shapes.String(90, height - 90, "Panel: {}".format(PanelNum), fontName = "Helvetica", fontSize = 20, textAnchor = 'middle'))

			font_size = 50
			text_width = width - 10
			name_width = stringWidth(name, "Helvetica", font_size)

			while name_width > text_width:
				font_size *= 0.8
				name_width = stringWidth(name, "Helvetica", font_size)
 
			s = shapes.String(width / 2.0, 70, FGN, textAnchor = "middle")
			s.fontName = "Helvetica"
			s.fontSize = font_size

			z = shapes.String(21 , 45, 'PN Built: _____  Initals:____')
			z.fontName = 'Helvetica'
			label.add(s)
			label.add(z)
 
		sheet = labels.Sheet(specs, write_sticker, border = True)
		sheet.add_label(FGN, Number_building)
		sheet.save('F:\\Harvest\\Python Print Document.pdf')
		#sheet.save('C:\\Users\\khanson\\Desktop\\Python Print Document.pdf')#change this to the path
		self.printfile()		

#code that creates a sticker to be put onto the substrate
############################################################################################################################################################################################################################################################

	def printfile(self):

		filename = 'F:\\Harvest\\Python Print Document.pdf'
		#filename = 'C:\\Users\\khanson\\Desktop\\Python Print Document.pdf'
		win32api.ShellExecute(0, 'print', filename, '/d:"%s"' % win32print.GetDefaultPrinter(), '.', 0)

#code that prints off the sticker that goes on the substrate
############################################################################################################################################################################################################################################################

	def pass_entry(self):

		passCheck = tk.Toplevel(root)
		passCheck.config(background = 'black')

		self.discrip = Label(passCheck)
		self.discrip.grid(row = 0, column = 0, pady = 5)
		self.discrip['text'] = 'Enter Password'
		self.discrip.config(font = ('Helvetica', 15), background = 'black', foreground = 'white')

		self.passresult = Label(passCheck)
		self.passresult.grid(row = 2, column = 0, pady = 5, padx = 10)
		self.passresult.config(background = 'black')

		self.passentry = Entry(passCheck, show = '*')
		self.passentry.grid(row = 1, column = 0, pady = 5, padx = 20)
		self.passentry.bind('<Key-Return>', self.pass_check)

#entry window to reveal visaid tab
############################################################################################################################################################################################################################################################

	def pass_check(self, event):

		password = self.passentry.get()

		if password == 'Eagle103':

			self.passresult['text'] = 'New VisAid Tab Opened on Main Window'
			self.passresult.config(font = ('Helvetica', 12), background = 'light green', foreground = 'black')
			self.show()

		else:
			self.passresult['text'] = 'Wrong Password Please Try Again'
			self.passresult.config(font = ('Helvetica', 15), background = 'red', foreground = 'white')

#password check
############################################################################################################################################################################################################################################################

	def Read_BoM(self):
		
		FGN = self.textentry.get()
		buildNum = int(self.NoB.get())

		
		wb = pyxl.load_workbook('F:\\Harvest\\Trial BoM\'s\\{}.xlsx'.format(bomdictionary[FGN]))
		ws = wb[FGN]

		value = 0
 
		for row in ws.iter_rows(min_col = 1, max_col = 1, min_row = 1, max_row = 50):
			for cell in row:
				if cell.value is None:
					last_row = value
				else:
					value = value + 1
					print(value)
 
		for row in ws.iter_rows(min_col = 3, max_col = 3, min_row = 3, max_row = last_row):
			for cell in row:
				quantity.append(int(cell.value)*buildNum)
 
 
		for row in ws.iter_rows(min_col = 2, max_col = 2, min_row = 3, max_row = last_row):
			for cell in row:
				part.append(cell.value)
 
 
		pTOq = [list(x) for x in zip(quantity, part)]

		l = listbox(root, height = 5)
		l.grid(column = 0, row = 0, sticky = (N, W, E, S))

		s = ttk.Scrollbar(root, orient = VERTICAL, command = l.yview)
		s.grid(column = 1, row = 0, sticky = (N, S))

		l['yscrollcommand'] = s.set 
		ttk.Sizegrip().grid(column = 1, row = 1, sticky = (S, E))
		root.grid_columnconfigure(0, weight = 1)
		root.grid_rowconfigure(0, weight = 1)

		for i in range (1, 101):
			l.insert('end', 'Line %d of 100' % i)

 
		print(pTOq)
		self.drawpdf()


#Reads the BoM and pulls out part list and calculates quantity based on build number
############################################################################################################################################################################################################################################################

	def addlight(self):

		FGN = self.FGNEntry.get()
		Lightguide = self.lghtEntry.get()

		if FGN in lghtguidedictionary:

			display = 'LightGuide for Finished Good Number Already exits as: {}'.format(lghtguidedictionary[FGN])
			self.job.config(font = ('Helvetica', 20), background = 'red')

		elif FGN == '':

			display = 'No FGN entry made Please input an entry.'
			self.job.config(font = ('Helvetica', 20), background = 'red')

		elif Lightguide == '':

			display = 'No LightGuide entry made Please input an entry.'
			self.job.config(font = ('Helvetica', 20), background = 'red')

		else:

			gather = open('lghtgid.json', 'w')
			lghtguidedictionary[self.FGNEntry.get()] = self.lghtEntry.get()
			gather.write(json.dumps(lghtguidedictionary))
			gather.close()
			display = 'Visual Aid: {} added as: {}'.format(FGN, lghtguidedictionary[FGN])
			self.job.config(font = ('Helvetica', 20), background = 'light green')

		self.job['text'] = display

#Adds LightGuide Program to dictionary
############################################################################################################################################################################################################################################################

	def addBoM(self):

		FGN = self.FGNEntry.get()
		BoM = self.BoMEntry.get()

		if FGN in bomdictionary:

			display = 'BoM for Finished Good Number Already exits as: {}'.format(bomdictionary[FGN])
			self.job.config(font = ('Helvetica', 20), background = 'red')

		elif FGN == '':

			display = 'No FGN entry made Please input an entry.'
			self.job.config(font = ('Helvetica', 20), background = 'red')

		elif BoM == '':

			display = 'No BoM entry made Please input an entry.'
			self.job.config(font = ('Helvetica', 20), background = 'red')

		else:

			gather = open('bom.json', 'w')
			bomdictionary[self.FGNEntry.get()] = self.BoMEntry.get()
			gather.write(json.dumps(bomdictionary))
			gather.close()
			display = 'BoM for FGN: {} added as: {}'.format(FGN, bomdictionary[FGN])
			self.job.config(font = ('Helvetica', 20), background = 'light green')

		self.job['text'] = display

#Adds Bom to dictionary
############################################################################################################################################################################################################################################################


root = tk.Tk()
app = Application(master = root)
app.mainloop()